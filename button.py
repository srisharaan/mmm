from tkinter import *
import openpyxl as xl

wb1=xl.load_workbook("book.xlsx")
ws1= wb1.worksheets[0]

def getdet():
    #wb1 = xl.load_workbook("book.xlsx")
    #ws1 = wb1.worksheets[0]
    global ws1
    global wb1
    root = Tk()
    root.geometry("600x250")
    root['background']='#99643A'
    #root.configure(bg='brown')
    root.title("MMM")
    def button2_click():
        root.destroy()
    def button1_click():
        #e1.delete(0,END)
        #e1.insert("saved")
        idd=e1.get()
        at=e2.get()
        tph=e3.get()
        no=e4.get()
        ws1.cell(row=1, column=1).value = 1
        ws1.cell(row=2, column=2).value = idd
        ws1.cell(row=3, column=2).value = at
        ws1.cell(row=4, column=2).value = tph
        ws1.cell(row=5, column=2).value = no
        wb1.save("book.xlsx")
        #mylabel=Label(root,text="saved")
        mylabel=Label(root,text="saved,click next to continue")
        mylabel.grid(row=9,column=0,columnspan=3,padx=10,pady=10)
        button_2 = Button(root, text="Next", padx=20, pady=10, command=button2_click)
        button_2.grid(row=11, column=1)



    mylabel1=Label(root,text="Accound SID:")
    mylabel1.grid(row=3,column=0)
    e1 = Entry(root, width=35, borderwidth=5)
    e1.grid(row=3, column=1,sticky="ew")
    

    mylabel2=Label(root,text="Auth Token:")
    mylabel2.grid(row=4,column=0)
    e2 = Entry(root, width=35, borderwidth=5)
    e2.grid(row=4, column=1,sticky="ew")

    mylabel3=Label(root,text="Twilio phone number:")
    mylabel3.grid(row=5,column=0)
    e3 = Entry(root, width=35, borderwidth=5)
    e3.grid(row=5, column=1,sticky="ew")

    mylabel4=Label(root,text="Your phone number(add country code)")
    mylabel4.grid(row=6,column=0)
    e4 = Entry(root, width=35, borderwidth=5)
    e4.grid(row=6, column=1,sticky="ew")


    button_1 = Button(root, text="Ok", padx=20, pady=10, command=button1_click)
    button_1.grid(row=10, column=1)

    root.mainloop()

#getdet()
