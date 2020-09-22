#from peru import getname
#from button import getdet
#from rece import *
from peru import getname
from button import getdet
import openpyxl as xl
import speech_recognition as sr
import re
import openpyxl as xl;
from twilio.rest import Client
from chumma import hi
import xlrd
from tkinter import *
import tkinter.font as font
from PIL import Image, ImageTk
import tkinter as tk


def hear(arr,j,k):
    #print(arr)
    slave=Tk()
    slave['background']='#99643A'
    slave.title("MMM")
    myFont = font.Font(family='Tw Cen MT', size=16)
    mylabel1=Label(slave,text="finding whether your hotword is being mentioned",bg='#99643A',fg='white')
    mylabel1.grid(row=3,column=0)
    slave.update()
    while(1):
        #print("first")
        i=0
        # obtain audio from the microphone
        r = sr.Recognizer()
        with sr.Microphone() as source:
            audio = r.listen(source)
            #print("about to start")
            try:
                #print("started")
                # for testing purposes, we're just using the default API key
                # to use another API key, use `r.recognize_google(audio, key="GOOGLE_SPEECH_RECOGNITION_API_KEY")`
                # instead of `r.recognize_google(audio)`
                a=r.recognize_google(audio)
                #slave.mainloop()
                #print(a)
                a=str(a)
                a=a.lower()
                #print(a)
                #print(type(a))
                flage=[]
                while(i<j):
                    jumma = arr[i]
                    kumma = str(jumma)
                    flage.append(kumma)
                    i=i+1
 
                #print("flage")
                #print(flage)
                temp=hi(flage,a)
                #print(match)
                    #print("insideuh")
                if (temp==1):
                    sms(flage,k)
                    #return 1
                    break


                       
            except sr.UnknownValueError:
                print("Could not understand audio")
                #mylabel1=Label(slave,text="listening",bg='#99643A',fg='white')
                #mylabel1.grid(row=3,column=0)
                #slave.update
            except sr.RequestError as e:
                print("Could not request results{0}".format(e))
                #mylabel1=Label(slave,text="listening",bg='#99643A',fg='white')
                #mylabel1.grid(row=3,column=0)
                #slave.update
            #slave.mainloop()
    slave.destroy()   
        
        



def sms(flage,k):
    wb1 = xl.load_workbook("book.xlsx")
    ws1 = wb1.worksheets[0]
    acc=ws1.cell(row = 2, column = 2).value
    uid=ws1.cell(row = 3, column = 2).value
    no=ws1.cell(row = 4, column = 2).value
    too=ws1.cell(row = 5, column = 2).value
    #too="+91"+too
    #print(too)
    #print(k)
    k=int(k)
    #wb1.save("book.xlsx")
    

    if(k==1):
        #mylabel=Label(win,text="calling",bg='#99643A',fg='white')
        #mylabel.grid(row=3,column=0)
        #win.update()
        client = Client(acc,uid)
        call = client.calls.create(
            url='https://demo.twilio.com/welcome/voice/',
            to=too,
            from_=no)
        

    if(k==2):
        #mylabel=Label(win,text="sms sent",bg='#99643A',fg='white')
        #mylabel.grid(row=3,column=0)
        # the following line needs your Twilio Account SID and Auth Token
        client = Client(acc,uid)
        #print(flage)
        msg="Hi,this message is from MMM because your hot word was mentioned."
        # change the "from_" number to your Twilio number and the "to" number
        # to the phone number you signed up for Twilio with, or upgrade your
        # account to send SMS to any phone number
        client.messages.create(to=too,from_=no,body=msg)
    if(k==3):
        #mylabel=Label(win,text="calling/sms sent",bg='#99643A',fg='white')
        #mylabel.grid(row=3,column=0)
        client = Client(acc,uid)
        msg="Hi,this message is from MMM because your hot word was mentioned."
        client.messages.create(to=too,from_=no,body=msg)
        call = client.calls.create(
            url='https://demo.twilio.com/welcome/voice/',
            to=too,
            from_=no)
def func(k):
    #print(k)
    
    
    wb2=xl.load_workbook("name.xlsx")
    ws2= wb2.worksheets[0]

    wb1 = xl.load_workbook("book.xlsx")
    ws1 = wb1.worksheets[0]

    
    d = ws1.cell(row = 1, column = 1).value
    d=int(d)
    #print(d)
    #print(type(d))
    c = ws2.cell(row = 1, column = 1).value
    c=int(c)
    #print(c)
    #print(type(c))
    if(d==0):
        #print("enter your twillio account details")
        getdet()
    #print("calling next")
    if(c==0):
        #print("should be opened")
        getname()

    j=ws2.cell(row = 2, column = 1).value
    #print(j)
    j=int(j)
    i=1
    arr=[]
    while(i<=j):
        temp=ws2.cell(row = i, column = 2).value
        arr.append(temp)
        i=i+1
    
    
    
    #print("starting to listen")
    k=int(k)
    while(1):
        hear(arr,j,k)
func(2)

