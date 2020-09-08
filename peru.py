from tkinter import *
import openpyxl as xl
from rece import *
i=1
wb2=xl.load_workbook("name.xlsx")
ws2= wb2.worksheets[0]
def getname():
    root = Tk()
    root.title("MMM")
    root.geometry("500x100")
    root['background']='#99643A'
    def button2_click():
        root.destroy()


    def button1_click():
        #wb2 = xl.load_workbook("name.xlsx")
        #ws2 = wb2.worksheets[0]
        global i
        global ws2
        global wb2
        # e.delete(0, END)
        current = e1.get()
        ws2.cell(row=i, column=2).value=current
        #i = i + 1
        ws2.cell(row=2, column=1).value=i
        ws2.cell(row=1, column=1).value = 1
        i = i + 1
        e1.delete(0,END)
        #e1.insert("saved")
        mylabel=Label(root,text="saved")
        mylabel.grid(row=9,column=0,columnspan=3,padx=10,pady=10)
        wb2.save("name.xlsx")




    mylabel1=Label(root,text="Enter the word/words for which you need to be notified.")
    mylabel1.grid(row=1,column=0)
    mylabel2=Label(root,text="For E.g:srisharaan,attendance")
    mylabel2.grid(row=1,column=1)
    e1 = Entry(root, width=30, borderwidth=1)
    e1.grid(row=2, column=0)
    button_1 = Button(root, text="Add word",command=button1_click)
    button_1.grid(row=3, column=0)

    button_2 = Button(root, text="finish",command=button2_click)
    button_2.grid(row=5, column=0)


    root.mainloop()
#getname()
