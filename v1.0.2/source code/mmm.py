from rece import hear

import openpyxl as xl; 

def name():
    wb2 = xl.load_workbook("name.xlsx")
    ws2 = wb2.worksheets[0]
    a=[]
    print("enter the trigger word for which you should be notified")
    print("enter all the ways your word(name) can be written")
    print("example: 1.keerthivaasan,2.kirtivasan,3.kerthivasan,4.kirtivaasan,5.keerthi,6.kirti,7.kriti,8.keethi,"
          "9.kithi,10.krithivaasan")
    j=1
    k=0
    while(k==0):
        print("enter the word")
        flag=input()
        #a.append(flag)
        #j=j+1
        print("do you want to continue (y/n)")
        y=input()
        if(y=='n'):
            k=1
        ws2.cell(row = j, column = 2).value =flag
        j=j+1
    j=j-1
    ws2.cell(row = 2, column = 1).value =j
    ws2.cell(row = 1, column = 1).value =1
    wb2.save("name.xlsx")
        


print("welcome")
wb1 = xl.load_workbook("book.xlsx")
ws1 = wb1.worksheets[0] 


c = ws1.cell(row = 1, column = 1).value
#print(c)
i=0
while(i==0):
    if (c==0):
        print("we require your twilio account credentials to send you sms")
        print("ACCOUNT SID:")
        sid=input()
        print("AUTH TOKEN:")
        tok=input()
        print("TWILIO PHONE NNUMBER:")
        no=input()
        print("The number in which you want to recieve the sms (add country code)")
        nu=input()
        print("Are the details given above are correct(y/n)")
        s=input()
        if(s=='y'):
            ws1.cell(row = 1, column = 1).value =1
            ws1.cell(row = 2, column = 2).value =sid
            ws1.cell(row = 3, column = 2).value =tok
            ws1.cell(row = 4, column = 2).value =no
            ws1.cell(row = 5, column = 2).value =nu
            wb1.save("book.xlsx")


            i=1
        
    else:
        print("lets get started!")
        i=1


wb2 = xl.load_workbook("name.xlsx")
ws2 = wb2.worksheets[0]
#print("o")

k = ws2.cell(row = 1, column = 1).value
print(k)
if(k==0):
    name()
    print("....")
else:
    print("...")



j=ws2.cell(row = 2, column = 1).value
print(j)
j=int(j)
i=1
arr=[]
while(i<=j):
    temp=ws2.cell(row = i, column = 2).value
    arr.append(temp)
    i=i+1
    
    
    
print("starting to listen")
while(1):
    
    hear(arr,j)



        
        
        

    


