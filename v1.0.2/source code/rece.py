import speech_recognition as sr
import re
import openpyxl as xl;
from twilio.rest import Client
from chumma import hello,hi
import xlrd


def hear(arr,j):
    i=0
    # obtain audio from the microphone
    r = sr.Recognizer()
    with sr.Microphone() as source:
        print("listening")
        audio = r.listen(source)
        #print("decoding")
        try:
            # for testing purposes, we're just using the default API key
            # to use another API key, use `r.recognize_google(audio, key="GOOGLE_SPEECH_RECOGNITION_API_KEY")`
            # instead of `r.recognize_google(audio)`
            a=r.recognize_google(audio)
            
            print(a)
            a=str(a)
            a=a.lower()
            print(a)
            #print(type(a))
            while(i<j):
                flage = arr[i]
                flage = str(flage)
                flage=flage.lower()
                
                #print("flage")
                #print(flage)
                temp=hi(flage,a)
                #print(match)
                i = i + 1
                if (temp==1):
                    sms(flage)
                    return 1
                    break

                
                
        except sr.UnknownValueError:
            print("Could not understand audio")
        except sr.RequestError as e:
            print("Could not request results{0}".format(e))



def sms(flage):
    wb1 = xl.load_workbook("book.xlsx")
    ws1 = wb1.worksheets[0]
    acc=ws1.cell(row = 2, column = 2).value
    uid=ws1.cell(row = 3, column = 2).value
    no=ws1.cell(row = 4, column = 2).value
    too=ws1.cell(row = 5, column = 2).value
    #too="+91"+too
    print(too)
    #wb1.save("book.xlsx")
    
    # the following line needs your Twilio Account SID and Auth Token
    client = Client(acc,uid)
    msg="Hi,this message is from MMM because the word "+flage+" was mentioned."
    # change the "from_" number to your Twilio number and the "to" number
    # to the phone number you signed up for Twilio with, or upgrade your
    # account to send SMS to any phone number
    client.messages.create(to=too,from_=no,body=msg)
    
        
        
    















